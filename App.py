from PlayerForm import PlayerForm
from MatchForm import MatchForm
from DataView import DataView
from ActionMenu import ActionMenu
from Model import Model
from typing import List
from datetime import datetime
import tkinter as tk
import sqlite3 as sql
import pandas as pd
import openpyxl
import ttkbootstrap as ttk
from ttkbootstrap.dialogs import Messagebox
from ttkbootstrap.constants import *
from ttkbootstrap.tooltip import ToolTip
from constants import *


class App(ttk.Window):
    def __init__(self, master=None, **kwargs):
        super().__init__(master, **kwargs)

        self.geometry(str(W_WIDTH) + "x" + str(W_HEIGHT))
        self.model = Model()
        self.conn = self.model.create_database()
        self.cur = self.conn.cursor()
        self.resizable(False, False)
        self.title("Rating Tracker")
        self.set_icon('images/Table-Tennis-1.png')

        self.create_widgets()
        self.actions = []
        self.actions_index = -1
        return

    def set_icon(self, icon_path: str) -> None:
        self.iconphoto(False, tk.PhotoImage(file=icon_path))
        return

    def create_widgets(self) -> None:
        self.player_form = PlayerForm(self, callback=lambda: self.callback_add_player
                                      (
                                          self.player_form.get_player().upper(),
                                          int(self.player_form.get_rating())
                                      ),
                                      text="Agregar jugador")
        self.match_form = MatchForm(self, text="Agregar resultado",
                                    callbacks=[
                                        lambda: self.callback_add_match(
                                            self.match_form.get_player_1(),
                                            self.match_form.get_player_2(),
                                            self.match_form.get_sets_1(),
                                            self.match_form.get_sets_2(),
                                            self.match_form.get_modalidad()
                                        ),
                                        self.callback_show_deltas
                                    ])
        self.data_view = DataView(self)
        self.btns_frame = ActionMenu(self,
                                     callbacks=[
                                         self.callback_undo,
                                         self.callback_redo,
                                         self.callback_export_excel
                                     ])

        self.match_form.grid(row=0, column=0, sticky="EW",
                             padx=(70, 0), pady=10)
        self.player_form.grid(
            row=1, column=0, sticky="EW", padx=(70, 0), pady=10)
        self.data_view.grid(row=2, column=0, sticky="EW",
                            padx=(70, 0), pady=10)
        self.btns_frame.grid(row=0, column=1, rowspan=3,
                             padx=5, pady=10, sticky="NS")

        self.match_form.update_player_opt(
            list(self.model.get_players()["Name"]))
        self.data_view.update_matches(self.model.get_matches())
        self.data_view.update_standings(self.model.get_standings())

        return

    def callback_add_player(self, jug: str, rating: int) -> None:
        try:
            self.model.add_player(jug, rating)
        except sql.Error as e:
            print(e)
            Messagebox.show_info(
                message="Fallo al agregar jugador", title="Error")
            return

        self.actions = self.actions[:self.actions_index + 1]
        self.actions.append(
            {
                "action": "add_player",
                "player": jug,
                "rating": rating
            }
        )
        self.actions_index += 1
        self.data_view.update_standings(self.model.get_standings())

        try:
            self.match_form.update_player_opt(
                list(self.model.get_players()["Name"]))
        except Exception as e:
            print(e)
            pass
        return

    def callback_show_deltas(self) -> None:
        jug1 = self.match_form.get_player_1()
        jug2 = self.match_form.get_player_2()

        players = self.model.get_players()[["Name", "Rating"]]
        players = players.loc[players["Name"].isin([jug1, jug2])]
        players = players.values.tolist()

        if len(players) != 2:
            print("Error: jugadores no encontrados")
            return
        deltas = None
        if players[0][0] == jug1:
            deltas = self.model.get_deltas(players[0][1], players[1][1],
                                           self.match_form.get_modalidad())
        else:
            deltas = self.model.get_deltas(players[1][1], players[0][1],
                                           self.match_form.get_modalidad())

        self.match_form.set_deltas(
            "+" + str(deltas[0][0]) + "/" + str(deltas[0][1]),
            "+" + str(deltas[1][0]) + "/" + str(deltas[1][1]))
        return

    def callback_add_match(self, jug1: str, jug2: str, sets_1: int, sets_2: int, modalidad: int) -> None:
        players = list(self.model.get_players()["Name"])

        max_sets = modalidad // 2 + 1
        error_msg = ""
        if jug1 == jug2:
            error_msg = "Error: jugadores iguales"
        elif (sets_1 == sets_2):
            error_msg = "Error: sets iguales"
        elif (sets_1 < 0 or sets_2 < 0):
            error_msg = "Error: sets negativos"
        elif ((not (jug1 in players)) or (not (jug2 in players))):
            error_msg = "Error: jugadores no encontrados"
        elif (sets_1 > max_sets or sets_2 > max_sets):
            error_msg = "Error: cantidad de sets mayor a la permitida"
        elif (sets_1 != max_sets and sets_2 != max_sets):
            error_msg = "Error: Tienen que llegar a " + str(max_sets) + " sets"
        if error_msg != "":
            Messagebox.show_info(title="Error", message=error_msg)
            return

        players = None
        try:
            players = self.model.get_players()[["PlayerId", "Name", "Rating"]]
        except sql.Error as e:
            print(e)
            return

        players = players.loc[players["Name"].isin([jug1, jug2])]
        players = players.values.tolist()
        jug1, jug2 = (players[0], players[1]) if players[0][1] == jug1 \
            else (players[1], players[0])
        time = datetime.now()

        try:
            match_id = self.model.add_match(
                jug1[0], jug2[0], jug1[2], jug2[2], sets_1, sets_2, time)
        except sql.Error as e:
            print(e)
            return

        deltas = self.model.get_deltas(jug1[2], jug2[2], modalidad)
        new_rating1, new_rating2 = (jug1[2] + deltas[0][0], jug2[2] + deltas[1][1]) \
            if sets_1 > sets_2 \
            else (jug1[2] + deltas[0][1], jug2[2] + deltas[1][0])
        
        try:
            self.model.update_player(["Rating"], [new_rating1], jug1[0])
            self.model.update_player(["Rating"], [new_rating2], jug2[0])
        except sql.Error as e:
            print(e)
            return

        self.data_view.update_standings(self.model.get_standings())
        self.data_view.update_matches(self.model.get_matches())
        self.actions = self.actions[:self.actions_index + 1]
        self.actions.append(
            {
                "action": "add_match",
                "match": match_id,
                "jug1": jug1[0],
                "jug2": jug2[0],
                "jug1_name": jug1[1],
                "jug2_name": jug2[1],
                "sets_1": sets_1,
                "sets_2": sets_2,
                "old_rating1": jug1[2],
                "old_rating2": jug2[2],
                "new_rating1": new_rating1,
                "new_rating2": new_rating2,
                "time": time
            }
        )
        self.actions_index += 1
        return 

    def callback_undo(self) -> None:
        if self.actions_index < 0:
            return
        last_action = self.actions[self.actions_index]
        if last_action["action"] == "add_match":
            try:
                self.cur.execute('''
                    SELECT Player1Id, Player1Rating, Player2Id, Player2Rating
                    FROM Match
                    WHERE MatchId = ?
                ''', (last_action["match"],))

                old_ratings = self.cur.fetchone()

                self.cur.execute('''
                    UPDATE Player
                    SET Rating = CASE PlayerId
                        WHEN ? THEN ?
                        WHEN ? THEN ?
                        ELSE Rating
                    END
                ''', (old_ratings[0], old_ratings[1], old_ratings[2], old_ratings[3]))

                self.cur.execute('''
                    DELETE FROM Match
                    WHERE MatchId = ?
                ''', (last_action["match"],))
            except sql as e:
                print(e)
            except Exception as e:
                print(e)
        elif last_action["action"] == "add_player":
            try:
                self.cur.execute('''
                    DELETE FROM Player
                    WHERE Name = ?
                ''', (last_action["player"],))
            except sql.Error as e:
                print(e)
            except Exception as e:
                print(e)

        self.conn.commit()
        self.data_view.update_standings(self.model.get_standings())
        self.data_view.update_matches(self.model.get_matches())
        self.actions_index -= 1
        return

    def callback_redo(self) -> None:
        if self.actions_index >= len(self.actions) - 1:
            return
        next_action = self.actions[self.actions_index + 1]

        if next_action["action"] == "add_player":
            self.cur.execute('''
                INSERT INTO Player (Name, Rating)
                VALUES (?, ?)
            ''', (next_action["player"], next_action["rating"]))
            self.conn.commit()
        elif next_action["action"] == "add_match":
            try:
                self.cur.execute('''
                    SELECT PlayerId
                    FROM Player
                    WHERE Name = ?
                ''', (next_action["jug1_name"],))
                jug1 = self.cur.fetchone()[0]
                self.cur.execute('''
                    SELECT PlayerId
                    FROM Player
                    WHERE Name = ?
                ''', (next_action["jug2_name"],))
                jug2 = self.cur.fetchone()[0]

                self.cur.execute('''
                    INSERT INTO Match (Player1Id, Player1Rating, Player2Id, Player2Rating, Player1Score, Player2Score, Date)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                ''', (jug1, next_action["old_rating1"], jug2, next_action["old_rating2"],
                      next_action["sets_1"], next_action["sets_2"], next_action["time"]))
                self.cur.execute('''
                    UPDATE Player
                    SET Rating = CASE PlayerId
                        WHEN ? THEN ?
                        WHEN ? THEN ?
                        ELSE Rating
                    END
                ''', (jug1, next_action["new_rating1"], jug2, next_action["new_rating2"]))
                self.conn.commit()
                self.actions[self.actions_index +
                             1]["match"] = self.cur.lastrowid

            except sql.Error as e:
                print(e.with_traceback())
                print(e)

        self.data_view.update_standings(self.model.get_standings())
        self.data_view.update_matches(self.model.get_matches())
        self.actions_index += 1
        return

    def callback_export_excel(self) -> None:
        wb = openpyxl.Workbook()

        ws = wb.create_sheet("Posiciones")
        ws.append(["Jugador", "Rating"])
        self.cur.execute('''
            SELECT Name, Rating
            FROM Player
            ORDER BY Rating DESC
        ''')
        for row in self.cur.fetchall():
            ws.append(row)

        ws = wb.create_sheet("Partidos")
        ws.append(["Jugador 1", "Resultado", "Jugador 2"])
        for match in self.model.get_matches():
            ws.append(match)

        del wb["Sheet"]
        wb.save("liga.xlsx")

        return
